import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os

os.makedirs('Docs', exist_ok=True)

def create_header_cell(sheet, row, col, text, width=None):
    cell = sheet.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if width:
        sheet.column_dimensions[get_column_letter(col)].width = width
    return cell

def create_section_header(sheet, row, col, text, span=1):
    cell = sheet.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    if span > 1:
        sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    return cell

int_wb = openpyxl.Workbook()
int_ws = int_wb.active
int_ws.title = 'システム構成'

int_ws.column_dimensions['A'].width = 20
int_ws.column_dimensions['B'].width = 60

int_ws.cell(row=1, column=1, value='devin_tokunou プロジェクト 内部設計書')
int_ws.merge_cells('A1:B1')
int_ws.cell(row=1, column=1).font = Font(bold=True, size=14)
int_ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

int_ws.cell(row=3, column=1, value='作成日')
int_ws.cell(row=3, column=2, value='2025年5月19日')
int_ws.cell(row=4, column=1, value='バージョン')
int_ws.cell(row=4, column=2, value='1.0')

create_section_header(int_ws, 6, 1, 'システムアーキテクチャ', 2)
int_ws.cell(row=7, column=1, value='アーキテクチャ')
int_ws.cell(row=7, column=2, value='FastAPIを使用したRESTful APIアーキテクチャ')
int_ws.cell(row=8, column=1, value='主要コンポーネント')
int_ws.cell(row=8, column=2, value='・FastAPIアプリケーション\n・ルーター（items, openai）\n・モジュール（item, openai）\n・スキーマ（item, openai）\n・ユーティリティ（logger）\n・ミドルウェア（CORS, リクエストログ）')
int_ws.cell(row=8, column=2).alignment = Alignment(wrap_text=True)
int_ws.row_dimensions[8].height = 90

create_section_header(int_ws, 10, 1, 'モジュール構成', 2)
int_ws.cell(row=11, column=1, value='app/main.py')
int_ws.cell(row=11, column=2, value='アプリケーションのエントリーポイント。FastAPIインスタンスの作成、ミドルウェアの設定、ルーターの登録、基本エンドポイントの定義、グローバル例外ハンドラの設定を行う。')
int_ws.cell(row=11, column=2).alignment = Alignment(wrap_text=True)
int_ws.row_dimensions[11].height = 45

int_ws.cell(row=12, column=1, value='app/routers/items.py')
int_ws.cell(row=12, column=2, value='アイテム管理のためのルーターとエンドポイント（GET /items/, GET /items/{item_id}, POST /items/）を定義。')
int_ws.cell(row=12, column=2).alignment = Alignment(wrap_text=True)
int_ws.row_dimensions[12].height = 45

int_ws.cell(row=13, column=1, value='app/routers/openai.py')
int_ws.cell(row=13, column=2, value='OpenAI APIとの連携のためのルーターとエンドポイント（POST /openai/chat/completions, GET /openai/models）を定義。')
int_ws.cell(row=13, column=2).alignment = Alignment(wrap_text=True)
int_ws.row_dimensions[13].height = 45

int_ws.cell(row=14, column=1, value='app/modules/item.py')
int_ws.cell(row=14, column=2, value='アイテム管理のビジネスロジックを実装。アイテムの取得、作成などの機能を提供。')
int_ws.cell(row=14, column=2).alignment = Alignment(wrap_text=True)
int_ws.row_dimensions[14].height = 45

int_ws.cell(row=15, column=1, value='app/modules/openai.py')
int_ws.cell(row=15, column=2, value='OpenAI APIとの連携ロジックを実装。チャット補完の生成、モデルリストの取得などの機能を提供。')
int_ws.cell(row=15, column=2).alignment = Alignment(wrap_text=True)
int_ws.row_dimensions[15].height = 45

int_ws.cell(row=16, column=1, value='app/schemas/item.py')
int_ws.cell(row=16, column=2, value='アイテム関連のデータモデル（Item, ItemResponse, ItemsResponse）を定義。')
int_ws.cell(row=16, column=2).alignment = Alignment(wrap_text=True)
int_ws.row_dimensions[16].height = 30

int_ws.cell(row=17, column=1, value='app/schemas/openai.py')
int_ws.cell(row=17, column=2, value='OpenAI API関連のデータモデル（ChatCompletionRequest）を定義。')
int_ws.cell(row=17, column=2).alignment = Alignment(wrap_text=True)
int_ws.row_dimensions[17].height = 30

int_ws.cell(row=18, column=1, value='app/utils/logger.py')
int_ws.cell(row=18, column=2, value='ロギング機能を提供。アプリケーションログ、APIログ、DBログの設定と初期化を行う。')
int_ws.cell(row=18, column=2).alignment = Alignment(wrap_text=True)
int_ws.row_dimensions[18].height = 30

create_section_header(int_ws, 20, 1, 'データフロー', 2)
int_ws.cell(row=21, column=1, value='アイテム一覧取得')
int_ws.cell(row=21, column=2, value='1. クライアントが GET /items/ にリクエスト\n2. リクエストログミドルウェアがリクエスト情報をログに記録\n3. items.pyのread_items関数が呼び出される\n4. item.pyのget_items関数が呼び出される\n5. fake_items_dbからすべてのアイテムが返される\n6. レスポンス情報がログに記録される')
int_ws.cell(row=21, column=2).alignment = Alignment(wrap_text=True)
int_ws.row_dimensions[21].height = 90

int_ws.cell(row=22, column=1, value='特定アイテム取得')
int_ws.cell(row=22, column=2, value='1. クライアントが GET /items/{item_id} にリクエスト\n2. リクエストログミドルウェアがリクエスト情報をログに記録\n3. items.pyのread_item関数が呼び出される\n4. item.pyのget_item関数が呼び出される\n5. item_idがfake_items_dbに存在するか確認\n6. 存在する場合はアイテムを返し、存在しない場合は404エラーを返す\n7. レスポンス情報がログに記録される')
int_ws.cell(row=22, column=2).alignment = Alignment(wrap_text=True)
int_ws.row_dimensions[22].height = 105

int_ws.cell(row=23, column=1, value='アイテム作成')
int_ws.cell(row=23, column=2, value='1. クライアントが POST /items/ にitem_idとnameを含めてリクエスト\n2. リクエストログミドルウェアがリクエスト情報をログに記録\n3. items.pyのcreate_item関数が呼び出される\n4. item.pyのcreate_item関数が呼び出される\n5. item_idがfake_items_dbに既に存在するか確認\n6. 存在しない場合は新しいアイテムを作成して返し、存在する場合は400エラーを返す\n7. レスポンス情報がログに記録される')
int_ws.cell(row=23, column=2).alignment = Alignment(wrap_text=True)
int_ws.row_dimensions[23].height = 105

int_ws.cell(row=24, column=1, value='OpenAIチャット補完')
int_ws.cell(row=24, column=2, value='1. クライアントが POST /openai/chat/completions にモデル名とメッセージを含めてリクエスト\n2. リクエストログミドルウェアがリクエスト情報をログに記録\n3. openai.pyのcreate_chat_completion関数が呼び出される\n4. modules/openai.pyのcreate_chat_completion関数が呼び出される\n5. OpenAIクライアントが初期化され、APIリクエストが送信される（開発環境ではモック応答）\n6. 結果がクライアントに返される\n7. レスポンス情報がログに記録される')
int_ws.cell(row=24, column=2).alignment = Alignment(wrap_text=True)
int_ws.row_dimensions[24].height = 105

create_section_header(int_ws, 26, 1, 'エラーハンドリング', 2)
int_ws.cell(row=27, column=1, value='アイテム未検出')
int_ws.cell(row=27, column=2, value='HTTPException(status_code=404, detail=f\"Item {item_id} not found\")を返す')
int_ws.cell(row=27, column=2).alignment = Alignment(wrap_text=True)

int_ws.cell(row=28, column=1, value='アイテム重複')
int_ws.cell(row=28, column=2, value='HTTPException(status_code=400, detail=f\"Item {item_id} already exists\")を返す')
int_ws.cell(row=28, column=2).alignment = Alignment(wrap_text=True)

int_ws.cell(row=29, column=1, value='OpenAI APIエラー')
int_ws.cell(row=29, column=2, value='HTTPException(status_code=500, detail=f\"OpenAI API error: {str(e)}\")を返す')
int_ws.cell(row=29, column=2).alignment = Alignment(wrap_text=True)

int_ws.cell(row=30, column=1, value='グローバル例外ハンドリング')
int_ws.cell(row=30, column=2, value='未処理の例外をキャッチし、ログに記録した後、HTTPException(status_code=500, detail=\"Internal Server Error\")を返す')
int_ws.cell(row=30, column=2).alignment = Alignment(wrap_text=True)
int_ws.row_dimensions[30].height = 45

create_section_header(int_ws, 32, 1, 'ロギング', 2)
int_ws.cell(row=33, column=1, value='ロガー種別')
int_ws.cell(row=33, column=2, value='・app_logger: アプリケーション全般のログ\n・api_logger: API呼び出しに関するログ\n・db_logger: データベース操作に関するログ')
int_ws.cell(row=33, column=2).alignment = Alignment(wrap_text=True)
int_ws.row_dimensions[33].height = 60

int_ws.cell(row=34, column=1, value='ログフォーマット')
int_ws.cell(row=34, column=2, value='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
int_ws.cell(row=34, column=2).alignment = Alignment(wrap_text=True)

int_ws.cell(row=35, column=1, value='リクエストログ')
int_ws.cell(row=35, column=2, value='リクエストIDを生成し、リクエスト開始時、完了時、エラー時にログを記録')
int_ws.cell(row=35, column=2).alignment = Alignment(wrap_text=True)

create_section_header(int_ws, 37, 1, '依存関係', 2)
int_ws.cell(row=38, column=1, value='外部ライブラリ')
int_ws.cell(row=38, column=2, value='・fastapi\n・uvicorn\n・starlette\n・pydantic\n・openai')
int_ws.cell(row=38, column=2).alignment = Alignment(wrap_text=True)
int_ws.row_dimensions[38].height = 75

int_wb.save('Docs/内部設計書.xlsx')
print('内部設計書.xlsx created successfully in the Docs folder.')
