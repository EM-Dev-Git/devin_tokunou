import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os

os.makedirs('Docs', exist_ok=True)

def create_header_cell(sheet, row, col, text, width=None):
    cell = sheet.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if width:
        sheet.column_dimensions[get_column_letter(col)].width = width
    return cell

def create_section_header(sheet, row, col, text, span=1):
    cell = sheet.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    if span > 1:
        sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    return cell

ext_wb = openpyxl.Workbook()
ext_ws = ext_wb.active
ext_ws.title = 'API仕様'

ext_ws.column_dimensions['A'].width = 20
ext_ws.column_dimensions['B'].width = 15
ext_ws.column_dimensions['C'].width = 15
ext_ws.column_dimensions['D'].width = 40

ext_ws.cell(row=1, column=1, value='devin_tokunou プロジェクト 外部設計書')
ext_ws.merge_cells('A1:D1')
ext_ws.cell(row=1, column=1).font = Font(bold=True, size=14)
ext_ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

ext_ws.cell(row=3, column=1, value='作成日')
ext_ws.cell(row=3, column=2, value='2025年5月19日')
ext_ws.cell(row=4, column=1, value='バージョン')
ext_ws.cell(row=4, column=2, value='1.0')

create_section_header(ext_ws, 6, 1, 'APIエンドポイント一覧', 4)
create_header_cell(ext_ws, 7, 1, 'エンドポイント', 20)
create_header_cell(ext_ws, 7, 2, 'メソッド', 15)
create_header_cell(ext_ws, 7, 3, 'パラメータ', 15)
create_header_cell(ext_ws, 7, 4, '説明', 40)

ext_ws.cell(row=8, column=1, value='/')
ext_ws.cell(row=8, column=2, value='GET')
ext_ws.cell(row=8, column=3, value='なし')
ext_ws.cell(row=8, column=4, value='ウェルカムメッセージを返す')

ext_ws.cell(row=9, column=1, value='/health')
ext_ws.cell(row=9, column=2, value='GET')
ext_ws.cell(row=9, column=3, value='なし')
ext_ws.cell(row=9, column=4, value='ヘルスチェック（アプリケーションの状態確認）')

ext_ws.cell(row=10, column=1, value='/items/')
ext_ws.cell(row=10, column=2, value='GET')
ext_ws.cell(row=10, column=3, value='なし')
ext_ws.cell(row=10, column=4, value='全アイテムのリストを返す')

ext_ws.cell(row=11, column=1, value='/items/{item_id}')
ext_ws.cell(row=11, column=2, value='GET')
ext_ws.cell(row=11, column=3, value='item_id: str')
ext_ws.cell(row=11, column=4, value='指定されたIDのアイテムを返す')

ext_ws.cell(row=12, column=1, value='/items/')
ext_ws.cell(row=12, column=2, value='POST')
ext_ws.cell(row=12, column=3, value='item_id: str\nname: str')
ext_ws.cell(row=12, column=3).alignment = Alignment(wrap_text=True)
ext_ws.cell(row=12, column=4, value='新しいアイテムを作成する')

ext_ws.cell(row=13, column=1, value='/openai/chat/completions')
ext_ws.cell(row=13, column=2, value='POST')
ext_ws.cell(row=13, column=3, value='model: str\nmessages: List[Dict]')
ext_ws.cell(row=13, column=3).alignment = Alignment(wrap_text=True)
ext_ws.cell(row=13, column=4, value='OpenAI APIを使用してチャット補完を生成する')

ext_ws.cell(row=14, column=1, value='/openai/models')
ext_ws.cell(row=14, column=2, value='GET')
ext_ws.cell(row=14, column=3, value='なし')
ext_ws.cell(row=14, column=4, value='利用可能なOpenAIモデルのリストを返す')

create_section_header(ext_ws, 16, 1, 'データモデル', 4)
create_header_cell(ext_ws, 17, 1, 'モデル名', 20)
create_header_cell(ext_ws, 17, 2, 'フィールド', 15)
create_header_cell(ext_ws, 17, 3, '型', 15)
create_header_cell(ext_ws, 17, 4, '説明', 40)

ext_ws.cell(row=18, column=1, value='Item')
ext_ws.cell(row=18, column=2, value='name')
ext_ws.cell(row=18, column=3, value='string')
ext_ws.cell(row=18, column=4, value='アイテムの名前')

ext_ws.cell(row=19, column=1, value='ItemResponse')
ext_ws.cell(row=19, column=2, value='item_id\nname')
ext_ws.cell(row=19, column=2).alignment = Alignment(wrap_text=True)
ext_ws.cell(row=19, column=3, value='string\nstring')
ext_ws.cell(row=19, column=3).alignment = Alignment(wrap_text=True)
ext_ws.cell(row=19, column=4, value='アイテムのID\nアイテムの名前')
ext_ws.cell(row=19, column=4).alignment = Alignment(wrap_text=True)

ext_ws.cell(row=20, column=1, value='ItemsResponse')
ext_ws.cell(row=20, column=2, value='items')
ext_ws.cell(row=20, column=3, value='Dict[str, Item]')
ext_ws.cell(row=20, column=4, value='アイテムIDをキーとするアイテムのディクショナリ')

ext_ws.cell(row=21, column=1, value='ChatCompletionRequest')
ext_ws.cell(row=21, column=2, value='model\nmessages')
ext_ws.cell(row=21, column=2).alignment = Alignment(wrap_text=True)
ext_ws.cell(row=21, column=3, value='string\nList[Dict]')
ext_ws.cell(row=21, column=3).alignment = Alignment(wrap_text=True)
ext_ws.cell(row=21, column=4, value='使用するモデル名\nチャットメッセージのリスト')
ext_ws.cell(row=21, column=4).alignment = Alignment(wrap_text=True)

create_section_header(ext_ws, 23, 1, 'レスポンス形式', 4)
create_header_cell(ext_ws, 24, 1, 'エンドポイント', 20)
create_header_cell(ext_ws, 24, 2, 'ステータスコード', 15)
create_header_cell(ext_ws, 24, 3, 'レスポンス形式', 15)
create_header_cell(ext_ws, 24, 4, '説明', 40)

ext_ws.cell(row=25, column=1, value='/')
ext_ws.cell(row=25, column=2, value='200')
ext_ws.cell(row=25, column=3, value='JSON')
ext_ws.cell(row=25, column=4, value='{\"message\": \"Welcome to devin_tokunou API\"}')

ext_ws.cell(row=26, column=1, value='/health')
ext_ws.cell(row=26, column=2, value='200')
ext_ws.cell(row=26, column=3, value='JSON')
ext_ws.cell(row=26, column=4, value='{\"status\": \"healthy\"}')

ext_ws.cell(row=27, column=1, value='/items/')
ext_ws.cell(row=27, column=2, value='200')
ext_ws.cell(row=27, column=3, value='JSON')
ext_ws.cell(row=27, column=4, value='アイテムのディクショナリ')

ext_ws.cell(row=28, column=1, value='/items/{item_id}')
ext_ws.cell(row=28, column=2, value='200/404')
ext_ws.cell(row=28, column=3, value='JSON')
ext_ws.cell(row=28, column=4, value='指定されたアイテム / エラーメッセージ')

ext_ws.cell(row=29, column=1, value='/items/')
ext_ws.cell(row=29, column=2, value='200/400')
ext_ws.cell(row=29, column=3, value='JSON')
ext_ws.cell(row=29, column=4, value='作成されたアイテム / エラーメッセージ')

ext_ws.cell(row=30, column=1, value='/openai/chat/completions')
ext_ws.cell(row=30, column=2, value='200/500')
ext_ws.cell(row=30, column=3, value='JSON')
ext_ws.cell(row=30, column=4, value='チャット補完結果 / エラーメッセージ')

ext_ws.cell(row=31, column=1, value='/openai/models')
ext_ws.cell(row=31, column=2, value='200/500')
ext_ws.cell(row=31, column=3, value='JSON')
ext_ws.cell(row=31, column=4, value='モデルリスト / エラーメッセージ')

create_section_header(ext_ws, 33, 1, 'エラーハンドリング', 4)
create_header_cell(ext_ws, 34, 1, 'エラー種別', 20)
create_header_cell(ext_ws, 34, 2, 'ステータスコード', 15)
create_header_cell(ext_ws, 34, 3, 'レスポンス形式', 15)
create_header_cell(ext_ws, 34, 4, '説明', 40)

ext_ws.cell(row=35, column=1, value='アイテム未検出')
ext_ws.cell(row=35, column=2, value='404')
ext_ws.cell(row=35, column=3, value='JSON')
ext_ws.cell(row=35, column=4, value='{\"detail\": \"Item {item_id} not found\"}')

ext_ws.cell(row=36, column=1, value='アイテム重複')
ext_ws.cell(row=36, column=2, value='400')
ext_ws.cell(row=36, column=3, value='JSON')
ext_ws.cell(row=36, column=4, value='{\"detail\": \"Item {item_id} already exists\"}')

ext_ws.cell(row=37, column=1, value='OpenAI APIエラー')
ext_ws.cell(row=37, column=2, value='500')
ext_ws.cell(row=37, column=3, value='JSON')
ext_ws.cell(row=37, column=4, value='{\"detail\": \"OpenAI API error: {error_message}\"}')

ext_ws.cell(row=38, column=1, value='内部サーバーエラー')
ext_ws.cell(row=38, column=2, value='500')
ext_ws.cell(row=38, column=3, value='JSON')
ext_ws.cell(row=38, column=4, value='{\"detail\": \"Internal Server Error\"}')

ext_wb.save('Docs/外部設計書.xlsx')
print('外部設計書.xlsx created successfully in the Docs folder.')
