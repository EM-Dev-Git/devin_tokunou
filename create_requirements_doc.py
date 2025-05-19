import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os

os.makedirs('Docs', exist_ok=True)

def create_header_cell(sheet, row, col, text, width=None):
    cell = sheet.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if width:
        sheet.column_dimensions[get_column_letter(col)].width = width
    return cell

def create_section_header(sheet, row, col, text, span=1):
    cell = sheet.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    if span > 1:
        sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    return cell

req_wb = openpyxl.Workbook()
req_ws = req_wb.active
req_ws.title = '要件定義書'

req_ws.column_dimensions['A'].width = 20
req_ws.column_dimensions['B'].width = 60

req_ws.cell(row=1, column=1, value='devin_tokunou プロジェクト 要件定義書')
req_ws.merge_cells('A1:B1')
req_ws.cell(row=1, column=1).font = Font(bold=True, size=14)
req_ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')

req_ws.cell(row=3, column=1, value='作成日')
req_ws.cell(row=3, column=2, value='2025年5月19日')
req_ws.cell(row=4, column=1, value='バージョン')
req_ws.cell(row=4, column=2, value='1.0')

create_section_header(req_ws, 6, 1, 'システム概要', 2)
req_ws.cell(row=7, column=1, value='システム名')
req_ws.cell(row=7, column=2, value='devin_tokunou API')
req_ws.cell(row=8, column=1, value='概要')
req_ws.cell(row=8, column=2, value='FastAPIを使用した基本的なRESTful APIアプリケーション')
req_ws.cell(row=9, column=1, value='目的')
req_ws.cell(row=9, column=2, value='アイテム管理とOpenAI APIを活用した機能を提供するAPIサービス')

create_section_header(req_ws, 11, 1, '機能要件', 2)
req_ws.cell(row=12, column=1, value='基本機能')
req_ws.cell(row=12, column=2, value='・ヘルスチェックエンドポイント\n・アイテム一覧の取得\n・特定アイテムの取得\n・新規アイテムの作成\n・OpenAI APIを使用した機能')
req_ws.cell(row=12, column=2).alignment = Alignment(wrap_text=True)
req_ws.row_dimensions[12].height = 75

req_ws.cell(row=13, column=1, value='ログ機能')
req_ws.cell(row=13, column=2, value='・構造化されたログ出力\n・リクエスト情報のログ記録')
req_ws.cell(row=13, column=2).alignment = Alignment(wrap_text=True)
req_ws.row_dimensions[13].height = 45

create_section_header(req_ws, 15, 1, '非機能要件', 2)
req_ws.cell(row=16, column=1, value='技術要件')
req_ws.cell(row=16, column=2, value='・Python 3.12以上\n・FastAPI\n・Uvicorn\n・OpenAI API')
req_ws.cell(row=16, column=2).alignment = Alignment(wrap_text=True)
req_ws.row_dimensions[16].height = 60

req_ws.cell(row=17, column=1, value='セキュリティ要件')
req_ws.cell(row=17, column=2, value='・CORSサポート\n・APIキー管理')
req_ws.cell(row=17, column=2).alignment = Alignment(wrap_text=True)

req_ws.cell(row=18, column=1, value='ドキュメント要件')
req_ws.cell(row=18, column=2, value='・Swagger UI（/docs）\n・ReDoc（/redoc）\n・システムアーキテクチャ図')
req_ws.cell(row=18, column=2).alignment = Alignment(wrap_text=True)
req_ws.row_dimensions[18].height = 45

req_wb.save('Docs/要件定義書.xlsx')
print('要件定義書.xlsx created successfully in the Docs folder.')
